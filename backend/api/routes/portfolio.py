from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from pydantic import BaseModel
from db.database import get_db
from db.models import Portfolio, PortfolioAsset, PortfolioOptimization
from api.auth import get_current_user, get_optional_user
from db.models import User
import base64
import re

router = APIRouter(prefix="/portfolio", tags=["portfolio"])


class PortfolioRequest(BaseModel):
    name: str = "Mi Portafolio"
    tickers: list[str]
    current_weights: dict[str, float] | None = None


class TickerListRequest(BaseModel):
    tickers: list[str]
    current_weights: dict[str, float] | None = None


@router.post("/optimize")
def optimize_portfolio(
    req: TickerListRequest,
    current_user: User = Depends(get_current_user),
):
    """Run portfolio optimization without saving. Quick analysis endpoint."""
    if len(req.tickers) < 2:
        raise HTTPException(status_code=400, detail="Se necesitan al menos 2 activos para optimizar")
    if len(req.tickers) > 20:
        raise HTTPException(status_code=400, detail="Máximo 20 activos por portafolio")

    from services.portfolio_optimizer import optimize_portfolio
    from services.claude_service import generate_portfolio_analysis

    result = optimize_portfolio(req.tickers, req.current_weights)
    ai = generate_portfolio_analysis(req.tickers, result, req.current_weights)

    return {**result, "ai_analysis": ai}


@router.post("/")
def create_portfolio(
    req: PortfolioRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Create and optimize a saved portfolio."""
    if len(req.tickers) < 2:
        raise HTTPException(status_code=400, detail="Se necesitan al menos 2 activos")

    from services.portfolio_optimizer import optimize_portfolio
    from services.claude_service import generate_portfolio_analysis

    portfolio = Portfolio(user_id=current_user.id, name=req.name)
    db.add(portfolio)
    db.flush()

    for ticker in req.tickers:
        weight = req.current_weights.get(ticker.upper()) if req.current_weights else None
        db.add(PortfolioAsset(portfolio_id=portfolio.id, ticker=ticker.upper(), current_weight=weight))

    opt_result = optimize_portfolio(req.tickers, req.current_weights)
    ai = generate_portfolio_analysis(req.tickers, opt_result, req.current_weights)

    for method in ["markowitz", "sharpe"]:
        m = opt_result.get(method, {})
        db.add(PortfolioOptimization(
            portfolio_id=portfolio.id,
            method=method,
            weights=m.get("weights"),
            expected_return=m.get("expected_return"),
            expected_volatility=m.get("volatility"),
            sharpe_ratio=m.get("sharpe"),
            bullets=ai.get("bullets"),
            efficient_frontier=opt_result.get("efficient_frontier") if method == "markowitz" else None,
        ))

    db.commit()
    db.refresh(portfolio)

    return {
        "portfolio_id": portfolio.id,
        "name": portfolio.name,
        **opt_result,
        "ai_analysis": ai,
    }


@router.get("/")
def list_portfolios(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    portfolios = db.query(Portfolio).filter(Portfolio.user_id == current_user.id).all()
    return [
        {
            "id": p.id,
            "name": p.name,
            "tickers": [a.ticker for a in p.assets],
            "created_at": p.created_at.isoformat() if p.created_at else None,
        }
        for p in portfolios
    ]


@router.get("/{portfolio_id}")
def get_portfolio(
    portfolio_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    portfolio = db.query(Portfolio).filter(
        Portfolio.id == portfolio_id,
        Portfolio.user_id == current_user.id,
    ).first()
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portafolio no encontrado")

    optimizations = db.query(PortfolioOptimization).filter(
        PortfolioOptimization.portfolio_id == portfolio_id
    ).all()

    return {
        "id": portfolio.id,
        "name": portfolio.name,
        "assets": [{"ticker": a.ticker, "current_weight": a.current_weight} for a in portfolio.assets],
        "optimizations": [
            {
                "method": o.method,
                "weights": o.weights,
                "expected_return": o.expected_return,
                "volatility": o.expected_volatility,
                "sharpe": o.sharpe_ratio,
                "bullets": o.bullets,
                "efficient_frontier": o.efficient_frontier,
            }
            for o in optimizations
        ],
        "created_at": portfolio.created_at.isoformat() if portfolio.created_at else None,
    }


@router.post("/extract-from-image")
async def extract_from_image(
    file: UploadFile = File(...),
    current_user: User = Depends(get_optional_user),
):
    """Use Claude Vision to extract tickers and weights from a portfolio screenshot or Excel."""
    content = await file.read()
    filename = file.filename or ''

    # Handle Excel files
    if filename.endswith(('.xlsx', '.xls', '.csv')):
        try:
            import io
            tickers = []
            weights = {}

            if filename.endswith('.csv'):
                import csv
                text = content.decode('utf-8', errors='ignore')
                reader = csv.DictReader(io.StringIO(text))
                rows = list(reader)
            else:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                ws = wb.active
                headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({headers[i]: str(v).strip() if v is not None else '' for i, v in enumerate(row)})

            ticker_keys = ['ticker', 'symbol', 'accion', 'simbolo', 'stock', 'code']
            weight_keys = ['weight', 'peso', 'porcentaje', '%', 'allocation', 'pct', 'participacion']

            ticker_col = next((h for h in (rows[0].keys() if rows else []) if any(k in h.lower() for k in ticker_keys)), None)
            weight_col = next((h for h in (rows[0].keys() if rows else []) if any(k in h.lower() for k in weight_keys)), None)

            for row in rows:
                t = (row.get(ticker_col) or '').strip().upper()
                if t and re.match(r'^[A-Z]{1,5}(-[A-Z])?$', t):
                    tickers.append(t)
                    if weight_col:
                        try:
                            w = float(str(row.get(weight_col, '')).replace('%', '').strip())
                            weights[t] = w / 100 if w > 1 else w
                        except Exception:
                            pass

            if tickers:
                return {'tickers': tickers, 'weights': weights if weights else None, 'source': 'excel'}
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error al leer el archivo: {str(e)}")

    # Handle image files — use Claude Vision
    import anthropic, os
    client = anthropic.Anthropic(api_key=os.getenv('ANTHROPIC_API_KEY'))
    mime = file.content_type or 'image/png'
    b64 = base64.standard_b64encode(content).decode()

    msg = client.messages.create(
        model='claude-haiku-4-5-20251001',
        max_tokens=500,
        messages=[{
            'role': 'user',
            'content': [
                {
                    'type': 'image',
                    'source': {'type': 'base64', 'media_type': mime, 'data': b64},
                },
                {
                    'type': 'text',
                    'text': (
                        'Extrae los tickers (símbolos bursátiles) y sus pesos/porcentajes de esta imagen de portafolio. '
                        'Responde SOLO con JSON en este formato exacto: '
                        '{"tickers": ["AAPL","MSFT"], "weights": {"AAPL": 0.6, "MSFT": 0.4}} '
                        'Si no hay pesos, pon weights como null. '
                        'Los tickers deben ser en mayúsculas. Solo incluye tickers válidos de bolsa (1-5 letras).'
                    )
                }
            ]
        }]
    )

    try:
        import json
        text = msg.content[0].text.strip()
        # Extract JSON from response
        match = re.search(r'\{.*\}', text, re.DOTALL)
        data = json.loads(match.group()) if match else {}
        return {**data, 'source': 'image'}
    except Exception:
        raise HTTPException(status_code=422, detail='No se pudieron extraer tickers de la imagen')


@router.delete("/{portfolio_id}")
def delete_portfolio(
    portfolio_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    portfolio = db.query(Portfolio).filter(
        Portfolio.id == portfolio_id,
        Portfolio.user_id == current_user.id,
    ).first()
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portafolio no encontrado")
    db.delete(portfolio)
    db.commit()
    return {"ok": True}
